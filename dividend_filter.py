"""
股息入账筛选工具
Filters rows where column '业务名称' == '股息入账' and saves them as a new sheet.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import re
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
except ImportError as e:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ─────────────────────────── constants ───────────────────────────
TARGET_COL  = "业务名称"
TARGET_VAL  = "股息入账"
SHEET_NAME  = "股息入账汇总"

BG_DARK     = "#0f0f1a"
BG_CARD     = "#1a1a2e"
BG_CARD2    = "#16213e"
ACCENT      = "#7c3aed"        # purple
ACCENT2     = "#6d28d9"
SUCCESS     = "#10b981"
WARNING     = "#f59e0b"
TEXT_PRI    = "#f1f5f9"
TEXT_SEC    = "#94a3b8"
TEXT_MUT    = "#475569"
BORDER_CLR  = "#2d2d4e"


# ─────────────────────────── helpers ─────────────────────────────
def make_border():
    thin = Side(style="thin", color="C0C0C0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header_cell(cell, col_idx):
    """Apply gradient-like header style to a cell."""
    colors = ["4F46E5", "5B21B6", "6D28D9", "7C3AED", "8B5CF6"]
    fill_color = colors[col_idx % len(colors)]
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.font = Font(bold=True, color="FFFFFF", name="微软雅黑", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()


def style_data_cell(cell, row_idx, is_numeric=False):
    even_fill = PatternFill("solid", fgColor="F8F7FF")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    cell.fill = even_fill if row_idx % 2 == 0 else odd_fill
    cell.font = Font(name="微软雅黑", size=10)
    cell.alignment = Alignment(
        horizontal="right" if is_numeric else "left",
        vertical="center",
        wrap_text=False,
    )
    cell.border = make_border()


def is_numeric_series(series):
    try:
        pd.to_numeric(series.dropna())
        return True
    except (ValueError, TypeError):
        return False


# ─────────────────────────── core logic ──────────────────────────
def _read_all_sheets(filepath: str, log_fn) -> tuple:
    """
    Try multiple strategies to read an Excel / text-as-xls file.
    Returns (xl_or_None, sheet_names, all_dfs_dict, is_html, is_xls).

    Strategy order:
      1. openpyxl        (.xlsx / .xlsm)
      2. xlrd            (true binary .xls)
      3. TSV / GBK text  (tab-separated GBK text saved as .xls — most Chinese brokers)
      4. HTML-as-xls     (HTML table saved as .xls)
    """
    ext = Path(filepath).suffix.lower()
    errors = []

    # ── Strategy 1: openpyxl ────────────────────────────────────
    if ext in (".xlsx", ".xlsm", ".xlam"):
        try:
            xl = pd.ExcelFile(filepath, engine="openpyxl")
            return xl, xl.sheet_names, None, False, False
        except Exception as e:
            errors.append(f"openpyxl: {e}")
            log_fn(f"⚠️  openpyxl 读取失败: {e}")

    # ── Strategy 2: xlrd (real binary .xls) ─────────────────────
    try:
        xl = pd.ExcelFile(filepath, engine="xlrd")
        return xl, xl.sheet_names, None, False, True
    except Exception as e:
        errors.append(f"xlrd: {e}")
        log_fn(f"⚠️  xlrd 读取失败: {e}")

    # ── Strategy 3: TSV / GBK text (most common Chinese broker export)
    log_fn("🔄 尝试以文本（TSV）格式解析…")
    for enc in ("gb18030", "gbk", "utf-8-sig", "utf-8"):
        try:
            df = pd.read_csv(filepath, sep="\t", encoding=enc, dtype=str,
                             engine="python", on_bad_lines="skip")
            df.columns = df.columns.astype(str).str.strip()
            # Sanity check: must have at least 2 columns with real content
            if df.shape[1] >= 2 and not df.empty:
                dfs = {"Sheet1": df}
                log_fn(f"✅ 成功以 TSV/{enc} 模式读取（共 {len(df)} 行，{len(df.columns)} 列）")
                return None, list(dfs.keys()), dfs, True, True
        except Exception as e:
            errors.append(f"TSV/{enc}: {e}")

    # ── Strategy 4: HTML-as-xls ──────────────────────────────────
    log_fn("🔄 尝试以 HTML 格式解析…")
    for enc in ("gbk", "gb18030", "utf-8"):
        try:
            tables = pd.read_html(filepath, encoding=enc, dtype=str)
            if tables:
                dfs = {f"Sheet{i+1}": t for i, t in enumerate(tables)}
                log_fn(f"✅ 成功以 HTML/{enc} 模式读取")
                return None, list(dfs.keys()), dfs, True, True
        except Exception as e:
            errors.append(f"HTML/{enc}: {e}")

    raise ValueError(
        "无法解析此文件，已尝试以下方式均失败：\n"
        + "\n".join(f"  • {e}" for e in errors)
        + "\n\n可能原因：\n"
        "  • 文件正被 Excel 等程序打开（请关闭后重试）\n"
        "  • 文件已损坏或格式不支持"
    )


def _load_inventory_map(log_fn):
    inv_map = {}
    possible_names = ["stock_name_code_exchange_inventory.xlsx", "stock_name_code_exchange_inventory.xls"]
    inv_path = None
    for name in possible_names:
        p = Path(__file__).parent / name
        if p.exists():
            inv_path = p
            break
            
    if not inv_path:
        log_fn("⚠️  未找到 stock_name_code_exchange_inventory.xlsx/xls 文件，将跳过备注更新。")
        return inv_map
        
    try:
        log_fn(f"📖 正在读取库存文件: {inv_path.name}")
        engine = "openpyxl" if inv_path.suffix.lower() == ".xlsx" else "xlrd"
        df_inv = pd.read_excel(str(inv_path), engine=engine)
        
        def clean_code(val):
            if pd.isna(val):
                return None
            val_str = str(val).strip()
            if val_str.startswith('="') and val_str.endswith('"'):
                val_str = val_str[2:-1]
            val_str = re.sub(r'^0+', '', val_str)
            if not val_str:
                return '0'
            return val_str

        for _, row in df_inv.iterrows():
            if '证券代码' not in row or '证券名称' not in row:
                continue
            code = clean_code(row['证券代码'])
            name = str(row['证券名称']).strip()
            latest_buy = row.get('latest_buy', '')
            if pd.isna(latest_buy):
                latest_buy = ''
            else:
                latest_buy = str(latest_buy).strip()
                
            if code:
                if code not in inv_map:
                    inv_map[code] = []
                inv_map[code].append((name, latest_buy))
                
        log_fn(f"✅ 成功加载库存映射，共 {len(inv_map)} 个证券代码。")
    except Exception as e:
        log_fn(f"⚠️  读取库存文件失败: {e}")
        
    return inv_map


def process_excel(filepath: str, progress_cb=None, log_cb=None):
    """
    Reads the Excel file, filters rows, writes a new sheet.
    Returns (matched_count, sheet_name, output_path) on success.

    For legacy .xls / HTML files the result is written to a new .xlsx
    file in the same directory (openpyxl cannot write .xls format).
    """
    def log(msg):
        if log_cb:
            log_cb(msg)

    def prog(val):
        if progress_cb:
            progress_cb(val)

    fpath = Path(filepath)
    log(f"📂 正在读取文件: {fpath.name}")
    prog(10)

    # ── Read all sheets (multi-strategy) ────────────────────────
    xl, sheet_names, html_dfs, is_html, is_xls = _read_all_sheets(filepath, log)
    if is_html:
        log("✅ 成功以 HTML 模式读取（国内券商格式）")
    log(f"📋 共发现 {len(sheet_names)} 个工作表: {', '.join(sheet_names)}")
    prog(20)

    # ── Search for target column across all sheets ───────────────
    matched_df  = None
    source_sheet = None
    for idx, sname in enumerate(sheet_names):
        if sname == SHEET_NAME:      # skip output sheet if re-running
            continue
        # Use pre-parsed dict for HTML mode, xl.parse() for real Excel
        if is_html:
            df = html_dfs[sname].copy()
            # Some HTML tables have unnamed header rows — promote first row
            if df.columns.astype(str).str.startswith("Unnamed").all():
                df.columns = df.iloc[0].astype(str).str.strip()
                df = df.iloc[1:].reset_index(drop=True)
            df.columns = df.columns.astype(str).str.strip()
        else:
            df = xl.parse(sname, dtype=str)
            df.columns = df.columns.astype(str).str.strip()

        if TARGET_COL in df.columns:
            filtered = df[df[TARGET_COL].str.strip() == TARGET_VAL].copy()
            if len(filtered):
                matched_df = filtered if matched_df is None else pd.concat(
                    [matched_df, filtered], ignore_index=True
                )
                source_sheet = sname
                log(f"✅ 在工作表 「{sname}」 找到 {len(filtered)} 条「{TARGET_VAL}」记录")
        prog(20 + int(30 * (idx + 1) / len(sheet_names)))

    if matched_df is None or len(matched_df) == 0:
        raise ValueError(
            f"未找到任何「{TARGET_COL}」列值为「{TARGET_VAL}」的行。\n"
            "请确认列名 and 数值是否正确。"
        )

    total_rows = len(matched_df)
    log(f"🔍 共筛选出 {total_rows} 条记录，准备写入新工作表…")
    prog(60)

    # ── Determine output path ────────────────────────────────────
    need_new_file = is_xls or is_html   # openpyxl cannot write .xls / HTML
    if need_new_file:
        out_path = fpath.with_name(fpath.stem + "_股息入账汇总.xlsx")
        log(f"ℹ️  将另存为新文件: {out_path.name}")
        from openpyxl import Workbook
        wb = Workbook()
        wb.remove(wb.active)           # remove default empty sheet
    else:
        out_path = fpath
        wb = load_workbook(filepath)
        # Remove existing output sheet to allow re-run
        if SHEET_NAME in wb.sheetnames:
            del wb[SHEET_NAME]
            log(f"♻️  已删除旧的「{SHEET_NAME}」工作表")

    # ── Build the new sheet ──────────────────────────────────────
    ws = wb.create_sheet(title=SHEET_NAME)
    prog(65)

    columns = list(matched_df.columns)

    # ── Load inventory map ───────────────────────────────────────
    inv_map = _load_inventory_map(log)

    code_idx = None
    name_idx = None
    memo_idx = None
    for ci, col_name in enumerate(columns):
        if col_name == '证券代码':
            code_idx = ci
        elif col_name == '证券名称':
            name_idx = ci
        elif col_name == '备注':
            memo_idx = ci

    # If '备注' is not in columns, add it to columns list
    if memo_idx is None:
        columns.append('备注')
        memo_idx = len(columns) - 1

    # Header row
    for ci, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        style_header_cell(cell, ci - 1)

    ws.row_dimensions[1].height = 28
    prog(70)

    # Data rows
    # Pre-build list of rows to write
    rows_to_write = []
    
    def clean_code(val):
        if pd.isna(val):
            return None
        val_str = str(val).strip()
        if val_str.startswith('="') and val_str.endswith('"'):
            val_str = val_str[2:-1]
        val_str = re.sub(r'^0+', '', val_str)
        if not val_str:
            return '0'
        return val_str

    for _, row in matched_df.iterrows():
        row_vals = list(row)
        # Ensure row_vals has enough elements if we appended '备注'
        while len(row_vals) < len(columns):
            row_vals.append("")
            
        # Update '备注' column if we found match
        if code_idx is not None and memo_idx is not None and inv_map:
            cell_code = row_vals[code_idx]
            cell_name = row_vals[name_idx] if name_idx is not None else ""
            
            if cell_code is not None and not pd.isna(cell_code):
                cleaned_ui_code = clean_code(cell_code)
                ui_name = str(cell_name).strip() if cell_name else ""
                
                if cleaned_ui_code in inv_map:
                    candidates = inv_map[cleaned_ui_code]
                    matched_latest_buy = None
                    if len(candidates) == 1:
                        matched_latest_buy = candidates[0][1]
                    else:
                        for name, lb in candidates:
                            if ui_name in name or name in ui_name:
                                matched_latest_buy = lb
                                break
                        if matched_latest_buy is None:
                            matched_latest_buy = candidates[0][1]
                            
                    if matched_latest_buy:
                        row_vals[memo_idx] = matched_latest_buy
                        
        rows_to_write.append(row_vals)

    numeric_flags = {
        ci: is_numeric_series(matched_df.iloc[:, ci - 1])
        for ci in range(1, len(matched_df.columns) + 1)
    }

    for ri, row_vals in enumerate(rows_to_write, start=2):
        for ci, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            style_data_cell(cell, ri, is_numeric=numeric_flags.get(ci, False))
        ws.row_dimensions[ri].height = 22
        if ri % 50 == 0:
            prog(70 + int(20 * ri / total_rows))

    prog(90)

    # Auto-width columns
    for ci, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(ci)
        # Compute max length using rows_to_write
        max_len = max(
            len(str(col_name)),
            max(len(str(r[ci - 1])) for r in rows_to_write) if rows_to_write else 0
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

    # Freeze top row & auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(str(out_path))
    prog(100)
    log(f"💾 已保存工作表「{SHEET_NAME}」→ {out_path.name}")

    return total_rows, SHEET_NAME, str(out_path)


# ─────────────────────────── GUI ─────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("股息入账筛选工具")
        self.geometry("780x600")
        self.resizable(True, True)
        self.minsize(640, 520)
        self.configure(bg=BG_DARK)

        self._filepath = tk.StringVar(value="")
        self._build_ui()
        self._center_window()

    # ── layout ───────────────────────────────────────────────────
    def _build_ui(self):
        # ── top title bar ────────────────────────────────────────
        title_frame = tk.Frame(self, bg=BG_DARK)
        title_frame.pack(fill="x", pady=(28, 0), padx=32)

        tk.Label(
            title_frame, text="📊", font=("Segoe UI Emoji", 28),
            bg=BG_DARK, fg=ACCENT
        ).pack(side="left")

        title_txt = tk.Frame(title_frame, bg=BG_DARK)
        title_txt.pack(side="left", padx=12)

        tk.Label(
            title_txt, text="股息入账筛选工具",
            font=("微软雅黑", 20, "bold"),
            bg=BG_DARK, fg=TEXT_PRI
        ).pack(anchor="w")

        tk.Label(
            title_txt,
            text=f"自动筛选「{TARGET_COL}」= 「{TARGET_VAL}」的行，写入新工作表",
            font=("微软雅黑", 10),
            bg=BG_DARK, fg=TEXT_SEC
        ).pack(anchor="w")

        # ── card: file picker ─────────────────────────────────────
        card1 = self._card(pady_top=24)

        tk.Label(
            card1, text="第 1 步  —  选择 Excel 文件",
            font=("微软雅黑", 11, "bold"),
            bg=BG_CARD, fg=ACCENT
        ).pack(anchor="w", pady=(0, 10))

        row = tk.Frame(card1, bg=BG_CARD)
        row.pack(fill="x")

        self._path_entry = tk.Entry(
            row, textvariable=self._filepath,
            font=("微软雅黑", 10), bg=BG_CARD2, fg=TEXT_PRI,
            relief="flat", bd=0,
            insertbackground=TEXT_PRI,
            highlightthickness=1, highlightcolor=ACCENT,
            highlightbackground=BORDER_CLR
        )
        self._path_entry.pack(side="left", fill="x", expand=True, ipady=8, ipadx=6)

        browse_btn = tk.Button(
            row, text="  浏览…  ",
            font=("微软雅黑", 10, "bold"),
            bg=ACCENT, fg="white", activebackground=ACCENT2,
            activeforeground="white", relief="flat", bd=0,
            cursor="hand2", padx=14, pady=8,
            command=self._browse
        )
        browse_btn.pack(side="left", padx=(10, 0))
        self._add_hover(browse_btn, ACCENT, ACCENT2)

        # ── card: options / info ──────────────────────────────────
        card2 = self._card(pady_top=14)

        tk.Label(
            card2, text="第 2 步  —  确认筛选条件",
            font=("微软雅黑", 11, "bold"),
            bg=BG_CARD, fg=ACCENT
        ).pack(anchor="w", pady=(0, 10))

        info_grid = tk.Frame(card2, bg=BG_CARD)
        info_grid.pack(fill="x")

        self._info_row(info_grid, 0, "筛选列名", TARGET_COL,  SUCCESS)
        self._info_row(info_grid, 1, "筛选值",   TARGET_VAL,  WARNING)
        self._info_row(info_grid, 2, "输出工作表名", SHEET_NAME, ACCENT)

        # ── run button ────────────────────────────────────────────
        run_frame = tk.Frame(self, bg=BG_DARK)
        run_frame.pack(pady=18)

        self._run_btn = tk.Button(
            run_frame, text="  ▶  开始筛选并写入  ",
            font=("微软雅黑", 12, "bold"),
            bg=SUCCESS, fg="white",
            activebackground="#059669", activeforeground="white",
            relief="flat", bd=0, cursor="hand2",
            padx=24, pady=12,
            command=self._start
        )
        self._run_btn.pack()
        self._add_hover(self._run_btn, SUCCESS, "#059669")

        # ── progress ──────────────────────────────────────────────
        prog_frame = tk.Frame(self, bg=BG_DARK)
        prog_frame.pack(fill="x", padx=32)

        self._prog_var = tk.IntVar(value=0)
        self._prog_bar = ttk.Progressbar(
            prog_frame, variable=self._prog_var,
            maximum=100, mode="determinate",
            style="Custom.Horizontal.TProgressbar"
        )
        self._prog_bar.pack(fill="x", pady=(0, 4))

        self._prog_label = tk.Label(
            prog_frame, text="", font=("微软雅黑", 9),
            bg=BG_DARK, fg=TEXT_SEC
        )
        self._prog_label.pack(anchor="e")

        # ── log ───────────────────────────────────────────────────
        log_frame = tk.Frame(self, bg=BG_DARK)
        log_frame.pack(fill="both", expand=True, padx=32, pady=(0, 24))

        log_header = tk.Label(
            log_frame, text="日志", font=("微软雅黑", 9, "bold"),
            bg=BG_DARK, fg=TEXT_MUT
        )
        log_header.pack(anchor="w", pady=(4, 2))

        log_inner = tk.Frame(log_frame, bg=BG_CARD2,
                             highlightthickness=1, highlightbackground=BORDER_CLR)
        log_inner.pack(fill="both", expand=True)

        self._log_text = tk.Text(
            log_inner, font=("Consolas", 9),
            bg=BG_CARD2, fg=TEXT_SEC,
            relief="flat", bd=0,
            state="disabled", wrap="word",
            insertbackground=TEXT_PRI
        )
        scrollbar = ttk.Scrollbar(log_inner, command=self._log_text.yview)
        self._log_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self._log_text.pack(fill="both", expand=True, padx=6, pady=6)

        # ── ttk style ─────────────────────────────────────────────
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor=BG_CARD2, background=ACCENT,
            lightcolor=ACCENT, darkcolor=ACCENT2,
            bordercolor=BORDER_CLR, thickness=8
        )

    # ── helper widgets ────────────────────────────────────────────
    def _card(self, pady_top=14):
        outer = tk.Frame(self, bg=BG_DARK)
        outer.pack(fill="x", padx=32, pady=(pady_top, 0))
        inner = tk.Frame(
            outer, bg=BG_CARD,
            highlightthickness=1, highlightbackground=BORDER_CLR
        )
        inner.pack(fill="x")
        pad = tk.Frame(inner, bg=BG_CARD)
        pad.pack(fill="x", padx=18, pady=14)
        return pad

    def _info_row(self, parent, row, label, value, color):
        tk.Label(
            parent, text=label + "：",
            font=("微软雅黑", 10), bg=BG_CARD, fg=TEXT_SEC
        ).grid(row=row, column=0, sticky="w", pady=3, padx=(0, 12))
        tk.Label(
            parent, text=value,
            font=("微软雅黑", 10, "bold"), bg=BG_CARD, fg=color
        ).grid(row=row, column=1, sticky="w", pady=3)

    def _add_hover(self, btn, normal_bg, hover_bg):
        btn.bind("<Enter>", lambda e: btn.configure(bg=hover_bg))
        btn.bind("<Leave>", lambda e: btn.configure(bg=normal_bg))

    def _center_window(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    # ── actions ───────────────────────────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[
                ("Excel 文件", "*.xlsx *.xlsm *.xls"),
                ("所有文件", "*.*")
            ]
        )
        if path:
            self._filepath.set(path)
            self._log_clear()
            self._log(f"📁 已选择文件: {Path(path).name}")

    def _start(self):
        path = self._filepath.get().strip()
        if not path:
            messagebox.showwarning("未选择文件", "请先选择一个 Excel 文件。")
            return
        if not os.path.isfile(path):
            messagebox.showerror("文件不存在", f"找不到文件:\n{path}")
            return

        self._run_btn.configure(state="disabled", text="  ⏳ 处理中…  ")
        self._prog_var.set(0)
        self._prog_label.configure(text="")
        self._log_clear()

        thread = threading.Thread(target=self._worker, args=(path,), daemon=True)
        thread.start()

    def _worker(self, path):
        try:
            count, sheet, out_path = process_excel(
                path,
                progress_cb=self._set_progress,
                log_cb=self._log
            )
            self.after(0, self._on_success, count, sheet, out_path)
        except Exception as exc:
            import traceback
            self.after(0, self._on_error, traceback.format_exc())

    def _on_success(self, count, sheet, out_path):
        self._run_btn.configure(state="normal", text="  ▶  开始筛选并写入  ")
        self._log(f"\n🎉 完成！共 {count} 行数据已写入工作表「{sheet}」")
        messagebox.showinfo(
            "处理完成",
            f"✅ 筛选完成！\n\n共找到 {count} 条「{TARGET_VAL}」记录\n"
            f"已写入工作表「{sheet}」\n\n输出文件: {Path(out_path).name}\n路径: {out_path}"
        )

    def _on_error(self, msg):
        self._run_btn.configure(state="normal", text="  ▶  开始筛选并写入  ")
        # Show a short friendly message in dialog but full trace in log
        first_line = msg.strip().splitlines()[-1] if msg.strip() else msg
        self._log(f"\n❌ 错误详情:\n{msg}")
        messagebox.showerror("处理失败", f"发生错误：\n\n{first_line}\n\n（详细信息见日志）")

    # ── log helpers ───────────────────────────────────────────────
    def _log(self, msg):
        def _do():
            self._log_text.configure(state="normal")
            self._log_text.insert("end", msg + "\n")
            self._log_text.see("end")
            self._log_text.configure(state="disabled")
        self.after(0, _do)

    def _log_clear(self):
        self._log_text.configure(state="normal")
        self._log_text.delete("1.0", "end")
        self._log_text.configure(state="disabled")

    def _set_progress(self, val):
        def _do():
            self._prog_var.set(val)
            self._prog_label.configure(text=f"{val}%")
        self.after(0, _do)


# ─────────────────────────── entry ───────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()
